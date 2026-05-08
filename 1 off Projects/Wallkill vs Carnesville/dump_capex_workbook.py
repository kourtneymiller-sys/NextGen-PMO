from pathlib import Path
import openpyxl

BASE = Path(__file__).parent
for path in [BASE / 'Wallkill CapEx Push.xlsx', BASE / 'Wallkill vs Carnesville Spend by FY and Timeline.xlsx']:
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    out.append(f'WORKBOOK: {path.name}')
    for ws in wb.worksheets:
        out.append(f'--- SHEET: {ws.title} rows={ws.max_row} cols={ws.max_column} ---')
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = []
            for c in row:
                vals.append('' if c is None else str(c))
            out.append(f'{i:03d}: ' + ' | '.join(vals))
    txt = path.with_suffix('.dump.txt')
    txt.write_text('\n'.join(out), encoding='utf-8')
    print(txt)
